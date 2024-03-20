from tkinter import *
import tkinter as tk
from tkinter import ttk, filedialog
from tkinter import messagebox
import sqlite3
from reportlab.lib.pagesizes import *
from reportlab.platypus import SimpleDocTemplate, Table, PageBreak
from reportlab.lib import colors
from tkinter import ttk
import os
from subprocess import call
from openpyxl import Workbook
import openpyxl

def update(rows):
    trv.delete(*trv.get_children())
    for i in rows:
        trv.insert('', 'end', values=i)


def search14():
    q2 = q.get()
    # q1 = q3.get()
    conn = sqlite3.connect("NASEOH.db")
    cursor = conn.cursor()
    query = "SELECT * FROM feedback WHERE cid LIKE'%" + q2 + "%'"
    cursor.execute(query)
    rows = cursor.fetchall()
    update(rows)


def search15():
    q2 = q1.get()
    # q1 = q3.get()
    conn = sqlite3.connect("NASEOH.db")
    cursor = conn.cursor()
    query = "SELECT * FROM feedback WHERE name LIKE'%" + q2 + "%'"
    cursor.execute(query)
    rows = cursor.fetchall()
    update(rows)


def getrow(event):
    rowid = trv.identify_row(event.y)  # not really required
    item = trv.item(trv.focus())
    t1.set(item['values'][0])
    t2.set(item['values'][1])
    t3.set(item['values'][2])
    t4.set(item['values'][3])
    t5.set(item['values'][4])


def update_candidate():
    cid = t1.get()
    name = t2.get()
    stufeed = t3.get()
    compfeed = t4.get()
    naseohfeed = t5.get()
    conn = sqlite3.connect("NASEOH.db")
    cursor = conn.cursor()

    if messagebox.askyesno("Confirm update", "Are you sure you want to update the data?"):
        query = "UPDATE feedback SET cid=?,name=?,stufeed=?,compfeed=?,naseohfeed=? WHERE cid=" + cid
        cursor.execute(query, (cid, name, stufeed, compfeed, naseohfeed))
        query1 = "SELECT * FROM feedback"
        cursor.execute(query1)
        conn.commit()
    else:
        return True


def delete_candidate():
    cid = t1.get()
    name = t2.get()
    conn = sqlite3.connect("NASEOH.db")
    cursor = conn.cursor()
    if messagebox.askyesno("Confirm delete", "Are you sure you want to delete the data?"):

        query = "DELETE FROM feedback WHERE cid=" + cid
        cursor.execute(query)
        query1 = "SELECT * FROM feedback"
        cursor.execute(query1)
        conn.commit()
    else:
        return True


def add_new():
    cid = t1.get()
    name = t2.get()
    stufeed = t3.get()
    compfeed = t4.get()
    naseohfeed = t5.get()
    conn = sqlite3.connect("NASEOH.db")
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT * from registration WHERE cid='" + cid + "'")
        if cursor.fetchone():

            global result
            result = 1

        else:
            result = 0

        if (result == 1):
            cursor.execute(
                "INSERT OR REPLACE INTO feedback VALUES('" + cid + "','" + name +"','" + stufeed + "','" + compfeed + "','" + naseohfeed + "')")
            conn.commit()
            #conn.close()
        else:
            messagebox.askretrycancel('error', "Candidate Id does not exists")
    except:
        # messagebox.askretrycancel('error',"Invalid cid")
        return True

root = Tk()
q6 = StringVar()
q4 = StringVar()
q5 = StringVar()
q1 = StringVar()
q3 = StringVar()
q = StringVar()
q2 = StringVar()
t1 = StringVar()
t2 = StringVar()
t3 = StringVar()
t4 = StringVar()
t5 = StringVar()

root.title("Search Feedback")
width= root.winfo_screenwidth()
height= root.winfo_screenheight()
root.geometry("%dx%d" % (width, height))

wrapper1 = LabelFrame(root, text="Feedback",font=('Arial',10,'bold'))
wrapper2 = LabelFrame(root, text="Search",font=('Arial',10,'bold'))
wrapper3 = LabelFrame(root, text="Update data",font=('Arial',10,'bold'))
wrapper1.configure(background="white")
wrapper2.configure(background="white")
wrapper3.configure(background="white")

wrapper1.pack(fill="both", expand="yes", padx=20, pady=10)
wrapper2.pack(fill="both", expand="yes", padx=20, pady=10)
wrapper3.pack(fill="both", expand="yes", padx=20, pady=10)

style = ttk.Style()
style.theme_use('alt')
style.configure("TCombobox", fieldbackground= "white", background= "white")
style.configure('TScrollbar' , background= "#329965")
style.configure('CustomScroll.Horizontal.TScrollbar',  arrowcolor='white', background= "#329965")
style.configure('CustomScroll.Vertical.TScrollbar', arrowcolor='white', background= "#329965")
x_scroll = ttk.Scrollbar(wrapper1, orient=HORIZONTAL)
x_scroll['style'] = "CustomScroll.Horizontal.TScrollbar"
x_scroll.pack(side=BOTTOM, fill=X)

y_scroll = ttk.Scrollbar(wrapper1, orient=VERTICAL)
y_scroll['style'] = "CustomScroll.Vertical.TScrollbar"
y_scroll.pack(side=RIGHT, fill=Y)

trv = ttk.Treeview(wrapper1, columns=(1, 2, 3, 4, 5), show="headings", height="6",xscrollcommand=x_scroll.set, yscrollcommand=y_scroll.set)
trv.pack(side = TOP, expand = True, fill = BOTH)

trv.heading(1, text="Candidate ID")
trv.heading(2, text="Candidate Name")
trv.heading(3, text="Student Feedback")
trv.heading(4, text="Company Feedback")
trv.heading(5, text="NASEOH Feedback")

trv.bind('<Double 1>', getrow)
conn = sqlite3.connect("NASEOH.db")
cursor = conn.cursor()
query = "SELECT * FROM Feedback ORDER BY cid desc"
cursor.execute(query)
rows = cursor.fetchall()
update(rows)

x_scroll.config(command=trv.xview)
y_scroll.config(command=trv.yview)


def get_treeview_data(trv):
    data = []
    for item in trv.get_children():
        values = trv.item(item, 'values')
        data.append(values)
    return data


def create_pdf(data):

    current_dir = filedialog.askdirectory(title="Select the current directory")

    if not current_dir:
        print("No directory selected.")
        exit()

    if not os.path.isdir(current_dir):
        print("Invalid directory path.")
        exit()

    filename = filedialog.asksaveasfilename(
        title="Save PDF file",
        initialdir=current_dir,
        defaultextension=".pdf",
        filetypes=[("PDF Files", "*.pdf")]
    )

    if not filename:
        print("No filename selected.")
        exit()

    doc = SimpleDocTemplate(filename, pagesize=A0)

    elements = []
    headers = ["Candidate ID","Candidate Name","Student Feedback","Company Feedback","NASEOH Feedback"]  # Replace with your column names
  
    data.insert(0, headers)


    table = Table(data, colWidths=None,rowHeights=75,  splitByRow=1,
                  repeatRows=0, repeatCols=0, rowSplitRange=None, spaceBefore=None,
                  spaceAfter=None, cornerRadii=None)


    style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 18),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 16),
        ('HALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),
        ('BOX', (0, 0), (-1, -1), 2, colors.black),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 15),
    ]
    table.setStyle(style)
    elements.append(table)
    doc.build(elements)
    print("PDF created successfully.")


def pdf():

    data = get_treeview_data(trv)
    create_pdf(data)

lbl14 = Label(wrapper2, text='Search by Candidate ID',font=('Arial',10,'bold'))
lbl14.grid(row=0, column=0, padx=5, pady=3)
ent14 = Entry(wrapper2, textvariable=q,highlightthickness=1,highlightbackground="black",highlightcolor="black")
ent14.grid(row=0, column=1, padx=5, pady=3)
btn14 = Button(wrapper2, text="Search", command=search14,font=('Arial',10,'bold'))
btn14.grid(row=0, column=2, padx=5, pady=3)
#
lbl15 = Label(wrapper2, text='Search by Candidate Name',font=('Arial',10,'bold'))
lbl15.grid(row=1, column=0, padx=5, pady=3)
ent15 = Entry(wrapper2, textvariable=q1,highlightthickness=1,highlightbackground="black",highlightcolor="black")
ent15.grid(row=1, column=1, padx=5, pady=3)
btn15 = Button(wrapper2, text="Search", command=search15,font=('Arial',10,'bold'))
btn15.grid(row=1, column=2, padx=5, pady=3)

# User data section
lbl1 = Label(wrapper3, text='Candidate ID',font=('Arial',10,'bold'))
lbl1.grid(row=0, column=0, padx=5, pady=3)
ent1 = Entry(wrapper3, textvariable=t1,highlightthickness=1,highlightbackground="black",highlightcolor="black")
ent1.grid(row=0, column=1, padx=5, pady=3)

lbl2 = Label(wrapper3, text='Candidate name',font=('Arial',10,'bold'))
lbl2.grid(row=0, column=3, padx=5, pady=3)
ent2 = Entry(wrapper3, textvariable=t2,highlightthickness=1,highlightbackground="black",highlightcolor="black")
ent2.grid(row=0, column=4, padx=5, pady=3)


def Ok():
    global myresult

    cid = ent1.get()
    name = ent2.get()

    conn = sqlite3.connect("NASEOH.db")
    cursor = conn.cursor()

    try:
        cursor.execute("SELECT cid,name from registration WHERE cid='" + cid + "'")
        myresult = cursor.fetchall()

        for x in myresult:
            print(x)
        ent2.delete(0,tk.END)
        ent2.insert(tk.END, x[1])

    except Exception as e:
       print(e)
       conn.rollback()
       conn.close()

tk.Button(wrapper3, text="Search Name", command=Ok,height = 1, width = 13,font=('Arial',10,'bold')).grid(row=0, column=2)

lbl3 = Label(wrapper3, text='Student Feedback',font=('Arial',10,'bold'))
lbl3.grid(row=1, column=0, padx=5, pady=3)
ent3 = Entry(wrapper3, textvariable=t3,highlightthickness=1,highlightbackground="black",highlightcolor="black")
ent3.grid(row=1, column=1, padx=5, pady=3)

lbl4 = Label(wrapper3, text='Company Feedback',font=('Arial',10,'bold'))
lbl4.grid(row=1, column=2, padx=5, pady=3)
ent4 = Entry(wrapper3, textvariable=t4,highlightthickness=1,highlightbackground="black",highlightcolor="black")
ent4.grid(row=1, column=3, padx=5, pady=3)

lbl5 = Label(wrapper3, text='NASEOH Feedback',font=('Arial',10,'bold'))
lbl5.grid(row=1, column=4, padx=5, pady=3)
ent5 = Entry(wrapper3, textvariable=t5,highlightthickness=1,highlightbackground="black",highlightcolor="black")
ent5.grid(row=1, column=5, padx=5, pady=3)

up_btn = Button(wrapper3, text="Update", command=update_candidate,font=('Arial',10,'bold'))
add_btn = Button(wrapper3, text="Add New", command=add_new,font=('Arial',10,'bold'))
delete_btn = Button(wrapper3, text="Delete Data", command=delete_candidate,font=('Arial',10,'bold'))
pdf_but = Button(wrapper1, text="PDF",command=pdf,font=('Arial',10,'bold'))
pdf_but.pack(side = LEFT, expand = True, fill = BOTH)

def search_a():
    #q2 = q5.get()
    # q1 = q3.get()
    query = "Select * FROM feedback ORDER by cid desc "
    cursor.execute(query)
    rows = cursor.fetchall()
    update(rows)

btna=Button(wrapper1,text="display all data",command=search_a,font=('Arial',10,'bold'))
btna.pack(side = LEFT, expand = True, fill = BOTH)

def get_treeview(trv):
    dat = []
    columns = []
    for column in trv["columns"]:
        columns.append(column)
    dat.append(columns)
    for item in trv.get_children():
        values = trv.item(item, 'values')
        dat.append(values)
    return dat

def create_excel(dat,filename):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    headers = ["Candidate ID","Candidate Name","Student Feedback","Company Feedback","NASEOH Feedback"]  # Replace with your column names

    # Add headers to the table
    dat.insert(0, headers)
    for row in dat:
        sheet.append(row)

    workbook.save(filename)
    print(f"Excel file '{filename}' created successfully.")


def excel():
    # Assuming your Treeview widget is called 'treeview'
    dat = get_treeview(trv)

    current_dir = filedialog.askdirectory(title="Select the current directory")
    if not current_dir:
        print("No directory selected.")
        return

    filename = filedialog.asksaveasfilename(
        title="Save Excel file",
        initialdir=current_dir,
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx")]
    )
    if not filename:
        print("No filename selected.")
        return

    create_excel(dat, filename)


excel_but = Button(wrapper1, text="Excel", command=excel, font=('Arial', 10, 'bold'))
excel_but.pack(side = LEFT, expand = True, fill = BOTH)

def upload_excel():
    # Open file dialog to select the Excel file
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    # Connect to the SQLite database
    conn = sqlite3.connect("NASEOH.db")  # Replace "your_database.db" with your SQLite database file path
    cursor = conn.cursor()
    # Read data from the Excel file
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    # Iterate over rows and insert data into the database
    for row in worksheet.iter_rows(values_only=True):
        # Replace column_index with the appropriate index for each column in the row
        query = "INSERT OR REPLACE INTO feedback (cid,name,stufeed,compfeed,naseohfeed) VALUES (?, ?, ?,?,?)"
        values = (row[0],row[1],row[2],row[3],row[4])
        cursor.execute(query,values)

    conn.commit()

upload = Button(wrapper1, text="Upload Excel", command=upload_excel, font=('Arial', 10, 'bold'))
upload.pack(side = LEFT, expand = True, fill = BOTH)

up_btn.grid(row=4, column=0, padx=5, pady=10)
add_btn.grid(row=4, column=1, padx=5, pady=10)
delete_btn.grid(row=4, column=2, padx=5, pady=10)

def nextpg():
    root.destroy()
    call(["python", "home.py"])
homepg = Button(wrapper3, text="HOMEPAGE", command=nextpg, font=('Arial', 10, 'bold')).grid(row=4, column=3, padx=5,pady=10)

root.mainloop()

