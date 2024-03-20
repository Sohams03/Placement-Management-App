from tkinter import *
import tkinter as tk
from tkinter import ttk, filedialog
from tkinter import messagebox
import sqlite3
from tkcalendar import *
from reportlab.lib.pagesizes import *
from reportlab.platypus import SimpleDocTemplate, Table, PageBreak
from reportlab.lib import colors
from tkinter import ttk
import os
from subprocess import call
from openpyxl import Workbook
import openpyxl

def update(rows):
    trv.delete(*trv.get_children())
    for i in rows:
        trv.insert('', 'end', values=i)
#


def search14():
    q2 = q.get()
    #q1 = q3.get()
    conn = sqlite3.connect("NASEOH.db")
    cursor = conn.cursor()
    query = "SELECT comid,comname,location,position,quali,age,vacancy,vacdate FROM vacancy WHERE comname LIKE'%"+q2+"%'"
    cursor.execute(query)
    rows = cursor.fetchall()
    update(rows)

def search15():
    q2 = q1.get()
    # q1 = q3.get()
    conn = sqlite3.connect("NASEOH.db")
    cursor = conn.cursor()
    query = "SELECT comid,comname,location,position,quali,age,vacancy,vacdate FROM vacancy WHERE location LIKE'%" + q2 + "%'"
    cursor.execute(query)
    rows = cursor.fetchall()
    update(rows)

def search16():
    q2 = q3.get()
    # q1 = q3.get()
    conn = sqlite3.connect("NASEOH.db")
    cursor = conn.cursor()
    query = "SELECT comid,comname,location,position,quali,age,vacancy,vacdate FROM vacancy WHERE position LIKE'%" + q2 + "%'"
    cursor.execute(query)
    rows = cursor.fetchall()
    update(rows)

def search17():
    q2 = q4.get()
    # q1 = q3.get()
    conn = sqlite3.connect("NASEOH.db")
    cursor = conn.cursor()
    query = "SELECT comid,comname,location,position,quali,age,vacancy,vacdate FROM vacancy WHERE quali LIKE'%" + q2 + "%'"
    cursor.execute(query)
    rows = cursor.fetchall()
    update(rows)

def search18():
    q2 = q5.get()
    # q1 = q3.get()
    conn = sqlite3.connect("NASEOH.db")
    cursor = conn.cursor()
    query = "SELECT comid,comname,location,position,quali,age,vacancy,vacdate FROM vacancy WHERE age LIKE'%" + q2 + "%'"
    cursor.execute(query)
    rows = cursor.fetchall()
    update(rows)

def search19():
    q2 = q6.get()
    # q1 = q3.get()
    conn = sqlite3.connect("NASEOH.db")
    cursor = conn.cursor()
    query = "SELECT comid,comname,location,position,quali,age,vacancy,vacdate FROM vacancy WHERE vacancy LIKE'%" + q2 + "%'"
    cursor.execute(query)
    rows = cursor.fetchall()
    update(rows)

def search20():
    q2 = q7.get()
    # q1 = q3.get()
    conn = sqlite3.connect("NASEOH.db")
    cursor = conn.cursor()
    query = "SELECT comid,comname,location,position,quali,age,vacancy,vacdate FROM vacdate WHERE vacancy LIKE'%" + q2 + "%'"
    cursor.execute(query)
    rows = cursor.fetchall()
    update(rows)

def getrow(event):
    rowid = trv.identify_row(event.y)  # not really required
    item = trv.item(trv.focus())
    t1.set(item['values'][0])
    t2.set(item['values'][1])
    t3.set(item['values'][2])
    t4.set(item['values'][3])
    t5.set(item['values'][4])
    t6.set(item['values'][5])
    t7.set(item['values'][6])
    t8.set(item['values'][7])


def update_candidate():
    comid = t1.get()
    comname = t2.get()
    location = t3.get()
    position = t4.get()
    quali=t5.get()
    age=t6.get()
    vacancy = t7.get()
    vacdate = t8.get()

    conn = sqlite3.connect("NASEOH.db")
    cursor = conn.cursor()



    if messagebox.askyesno("Confirm update", "Are you sure you want to update the data?"):
        query = "UPDATE vacancy SET comid=?,comname=?,location=?,position=?,quali=?,age=?,vacancy=? ,vacdate=?WHERE comid=" + comid
        cursor.execute(query,(comid,comname,location,position,quali,age,vacancy,vacdate))
        query1 = "SELECT comid,comname,location,position,quali,age,vacancy,vacdate FROM vacancy"
        cursor.execute(query1)
        conn.commit()

    else:
        return True


def delete_candidate():
    comid = t1.get()
    conn = sqlite3.connect("NASEOH.db")
    cursor = conn.cursor()
    if messagebox.askyesno("Confirm delete", "Are you sure you want to delete the data?"):

        query = "DELETE FROM vacancy WHERE comid=" + comid
        cursor.execute(query)
        query1 = "SELECT comid,comname,location,position,quali,age,vacancy,vacdate FROM vacancy"
        cursor.execute(query1)
        conn.commit()

    else:
        return True


def add_new():
    comid = t1.get()
    comname = t2.get()
    location = t3.get()
    position = t4.get()
    quali=t5.get()
    age=t6.get()
    vacancy = t7.get()
    vacdate = t8.get()

    conn = sqlite3.connect("NASEOH.db")

    try:
        cursor = conn.cursor()
        cursor.execute("SELECT * from regiscomp WHERE comid='" + comid + "'")
        if cursor.fetchone():

            global result
            result = 1

        else:
            result = 0

        if (result == 1):

            cursor.execute(
                "INSERT OR REPLACE INTO vacancy VALUES('" + comid + "','" + comname +"','" + location + "','" + position +"','" + quali +"','" + age + "','" + vacancy + "','" + vacdate + "')")
            conn.commit()
            #conn.close()
        else:
            messagebox.askretrycancel('error', "Company Id does not exists")
    except:
        return True



def pick_date(event):
    global cal ,date_window

    date_window = Toplevel()
    date_window.grab_set()
    date_window.title('Choose Date')
    date_window.geometry('250x250')
    cal = Calendar(date_window, selectmode='day', date_pattern='dd/mm/y')
    cal.place(x=0,y=0)

    sub_btn = Button(date_window,text='Submit',command=grab_date)
    sub_btn.place(x=80,y=190)
def grab_date():
    ent20.delete(0,END)
    ent20.insert(0,cal.get_date())
    date_window.destroy()

def pick_date1(event):
    global cal ,date_window

    date_window = Toplevel()
    date_window.grab_set()
    date_window.title('Choose Date')
    date_window.geometry('250x250')
    cal = Calendar(date_window, selectmode='day', date_pattern='dd/mm/y')
    cal.place(x=0,y=0)

    sub_btn = Button(date_window,text='Submit',command=grab_date1)
    sub_btn.place(x=80,y=190)
def grab_date1():
    ent6.delete(0,END)
    ent6.insert(0,cal.get_date())
    date_window.destroy()


root = Tk()
q7 = StringVar()
q6 = StringVar()
q4 = StringVar()
q5 = StringVar()
q1 = StringVar()
q3 = StringVar()
q = StringVar()
q2 = StringVar()
t1 = StringVar()
t2 = StringVar()
t3 = StringVar()
t4 = StringVar()
t5=StringVar()
t6=StringVar()
t7=StringVar()
t8=StringVar()

root.title("Search Vacancy")
width= root.winfo_screenwidth()
height= root.winfo_screenheight()
root.geometry("%dx%d" % (width, height))

wrapper1 = LabelFrame(root, text="Vacancy",font=('Arial',10,'bold'))
wrapper2 = LabelFrame(root, text="Search",font=('Arial',10,'bold'))
wrapper3 = LabelFrame(root, text="Update data",font=('Arial',10,'bold'))
wrapper1.configure(background="white")
wrapper2.configure(background="white")
wrapper3.configure(background="white")

wrapper1.pack(fill="both", expand="yes", padx=20, pady=10)
wrapper2.pack(fill="both", expand="yes", padx=20, pady=10)
wrapper3.pack(fill="both", expand="yes", padx=20, pady=10)
style = ttk.Style()
style.theme_use('alt')
style.configure("TCombobox", fieldbackground= "white", background= "white")
style.configure('TScrollbar' , background= "#329965")
style.configure('CustomScroll.Horizontal.TScrollbar',  arrowcolor='white', background= "#329965")
style.configure('CustomScroll.Vertical.TScrollbar', arrowcolor='white', background= "#329965")
x_scroll = ttk.Scrollbar(wrapper1, orient=HORIZONTAL)
x_scroll['style'] = "CustomScroll.Horizontal.TScrollbar"
x_scroll.pack(side=BOTTOM, fill=X)

y_scroll = ttk.Scrollbar(wrapper1, orient=VERTICAL)
y_scroll['style'] = "CustomScroll.Vertical.TScrollbar"
y_scroll.pack(side=RIGHT, fill=Y)
trv = ttk.Treeview(wrapper1, columns=(1, 2, 3, 4,5,6,7,8), show="headings", height="7",xscrollcommand=x_scroll.set, yscrollcommand=y_scroll.set)
trv.pack(side = TOP, expand = True, fill = BOTH)

trv.heading(1, text="Company ID")
trv.heading(2, text="Company Name")
trv.heading(3, text="Company Location")
trv.heading(4, text="Position")
trv.heading(5, text="Qualification")
trv.heading(6, text="Age")
trv.heading(7, text="Vacancy")
trv.heading(8, text="Vacancy Date")

trv.bind('<Double 1>', getrow)
conn = sqlite3.connect("NASEOH.db")
cursor = conn.cursor()
query = "SELECT  comid,comname,location,position,quali,age,vacancy,vacdate FROM vacancy"
cursor.execute(query)
rows = cursor.fetchall()
update(rows)
x_scroll.config(command=trv.xview)
y_scroll.config(command=trv.yview)

def get_treeview_data(trv):
    data = []
    for item in trv.get_children():
        values = trv.item(item, 'values')
        data.append(values)
    return data


def create_pdf(data):

    current_dir = filedialog.askdirectory(title="Select the current directory")

    if not current_dir:
        print("No directory selected.")
        exit()

    if not os.path.isdir(current_dir):
        print("Invalid directory path.")
        exit()
    filename = filedialog.asksaveasfilename(
        title="Save PDF file",
        initialdir=current_dir,
        defaultextension=".pdf",
        filetypes=[("PDF Files", "*.pdf")]
    )

    if not filename:
        print("No filename selected.")
        exit()

    doc = SimpleDocTemplate(filename, pagesize=A2)

    elements = []
    headers = ["Company ID","Company Name","Company Location","Position","Qualification","Age Limit","Vacancy Date"]  # Replace with your column names

    data.insert(0, headers)

    table = Table(data, colWidths=None,rowHeights=75,  splitByRow=1,
                  repeatRows=0, repeatCols=0, rowSplitRange=None, spaceBefore=None,
                  spaceAfter=None, cornerRadii=None)


    style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 18),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 16),
        ('HALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),
        ('BOX', (0, 0), (-1, -1), 2, colors.black),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 15),
    ]
    table.setStyle(style)
    elements.append(table)
    doc.build(elements)
    print("PDF created successfully.")


def pdf():
    data = get_treeview_data(trv)
    create_pdf(data)


lbl14=Label(wrapper2,text='Search by Company name',font=('Arial',10,'bold'))
lbl14.grid(row=0,column=0,padx=5,pady=3)
ent14=Entry(wrapper2,textvariable=q,highlightthickness=1,highlightbackground="black",highlightcolor="black")
ent14.grid(row=0,column=1,padx=5,pady=3)
btn14=Button(wrapper2,text="Search",command=search14,font=('Arial',10,'bold'))
btn14.grid(row=0,column=2,padx=5,pady=3)
#
lbl15=Label(wrapper2,text='Search by Location',font=('Arial',10,'bold'))
lbl15.grid(row=1,column=0,padx=5,pady=3)
ent15=Entry(wrapper2,textvariable=q1,highlightthickness=1,highlightbackground="black",highlightcolor="black")
ent15.grid(row=1,column=1,padx=5,pady=3)
btn15=Button(wrapper2,text="Search",command=search15,font=('Arial',10,'bold'))
btn15.grid(row=1,column=2,padx=5,pady=3)

lbl16=Label(wrapper2,text='Search by Position',font=('Arial',10,'bold'))
lbl16.grid(row=2,column=0,padx=5,pady=3)
ent16=Entry(wrapper2,textvariable=q3,highlightthickness=1,highlightbackground="black",highlightcolor="black")
ent16.grid(row=2,column=1,padx=5,pady=3)
btn16=Button(wrapper2,text="Search",command=search16,font=('Arial',10,'bold'))
btn16.grid(row=2,column=2,padx=5,pady=16)

lbl17=Label(wrapper2,text='Search by Qualification',font=('Arial',10,'bold'))
lbl17.grid(row=0,column=3,padx=5,pady=3)
ent17=Entry(wrapper2,textvariable=q4,highlightthickness=1,highlightbackground="black",highlightcolor="black")
ent17.grid(row=0,column=4,padx=5,pady=3)
btn17=Button(wrapper2,text="Search",command=search17,font=('Arial',10,'bold'))
btn17.grid(row=0,column=5,padx=5,pady=16)

lbl18=Label(wrapper2,text='Search by Age',font=('Arial',10,'bold'))
lbl18.grid(row=1,column=3,padx=5,pady=3)
ent18=Entry(wrapper2,textvariable=q5,highlightthickness=1,highlightbackground="black",highlightcolor="black")
ent18.grid(row=1,column=4,padx=5,pady=3)
btn18=Button(wrapper2,text="Search",command=search18,font=('Arial',10,'bold'))
btn18.grid(row=1,column=5,padx=5,pady=16)

lbl19=Label(wrapper2,text='Search by Vacancy',font=('Arial',10,'bold'))
lbl19.grid(row=2,column=3,padx=5,pady=3)
ent19=Entry(wrapper2,textvariable=q6,highlightthickness=1,highlightbackground="black",highlightcolor="black")
ent19.grid(row=2,column=4,padx=5,pady=3)
btn19=Button(wrapper2,text="Search",command=search19,font=('Arial',10,'bold'))
btn19.grid(row=2,column=5,padx=5,pady=16)

lbl20= Label(wrapper2,text='Date',font=('Arial',10,'bold'))
lbl20.grid(row=0, column=6, padx=5, pady=3)
ent20= Entry(wrapper2,textvariable=q7,highlightthickness=1,highlightbackground="black",highlightcolor="black")
ent20.grid(row=0, column=7, padx=5, pady=3)
ent20.insert(0,"dd/mm/yyyy")
ent20.bind("<1>",pick_date)
btn20=Button(wrapper2,text="Search",command=search20,font=('Arial',10,'bold'))
btn20.grid(row=0,column=8,padx=5,pady=16)

# User data section
lbl1 = Label(wrapper3, text='Company ID',font=('Arial',10,'bold'))
lbl1.grid(row=0, column=0, padx=5, pady=3)
ent1 = Entry(wrapper3, textvariable=t1,highlightthickness=1,highlightbackground="black",highlightcolor="black")
ent1.grid(row=0, column=1, padx=5, pady=3)

lbl2 = Label(wrapper3, text='Company name',font=('Arial',10,'bold'))
lbl2.grid(row=1, column=0, padx=5, pady=3)
ent2 = Entry(wrapper3, textvariable=t2,highlightthickness=1,highlightbackground="black",highlightcolor="black")
ent2.grid(row=1, column=1, padx=5, pady=3)

lbl3 = Label(wrapper3, text='Location',font=('Arial',10,'bold'))
lbl3.grid(row=2, column=0, padx=5, pady=3)
ent3 = Entry(wrapper3, textvariable=t3,highlightthickness=1,highlightbackground="black",highlightcolor="black")
ent3.grid(row=2, column=1, padx=5, pady=3)

lbl4 = Label(wrapper3, text='Position',font=('Arial',10,'bold'))
lbl4.grid(row=3, column=0, padx=5, pady=3)
ent4 = Entry(wrapper3, textvariable=t4,highlightthickness=1,highlightbackground="black",highlightcolor="black")
ent4.grid(row=3, column=1, padx=5, pady=3)

lbl5 = Label(wrapper3, text='Qualification',font=('Arial',10,'bold'))
lbl5.grid(row=0, column=2, padx=5, pady=3)
ent5 = Entry(wrapper3, textvariable=t5,highlightthickness=1,highlightbackground="black",highlightcolor="black")
ent5.grid(row=0, column=3, padx=5, pady=3)

lbl6 = Label(wrapper3, text='Age',font=('Arial',10,'bold'))
lbl6.grid(row=1, column=2, padx=5, pady=3)
ent6 = Entry(wrapper3, textvariable=t6,highlightthickness=1,highlightbackground="black",highlightcolor="black")
ent6.grid(row=1, column=3, padx=5, pady=3)

lbl6 = Label(wrapper3, text='Vacancy',font=('Arial',10,'bold'))
lbl6.grid(row=2, column=2, padx=5, pady=3)
ent6 = Entry(wrapper3, textvariable=t7,highlightthickness=1,highlightbackground="black",highlightcolor="black")
ent6.grid(row=2, column=3, padx=5, pady=3)

lbl6 = Label(wrapper3, text='Vacancy Date',font=('Arial',10,'bold'))
lbl6.grid(row=3, column=2, padx=5, pady=3)
ent6 = Entry(wrapper3, textvariable=t8,highlightthickness=1,highlightbackground="black",highlightcolor="black")
ent6.grid(row=3, column=3, padx=5, pady=3)
ent6.insert(0,"dd/mm/yyyy")
ent6.bind("<1>",pick_date1)


up_btn = Button(wrapper3, text="Update", command=update_candidate,font=('Arial',10,'bold'))
add_btn = Button(wrapper3, text="Add New", command=add_new,font=('Arial',10,'bold'))
delete_btn = Button(wrapper3, text="Delete Data", command=delete_candidate,font=('Arial',10,'bold'))
pdf_but = Button(wrapper1, text="PDF",command=pdf,font=('Arial',10,'bold'))
pdf_but.pack(side = LEFT, expand = True, fill = BOTH)

def search_a():
    #q2 = q5.get()
    # q1 = q3.get()
    query = "Select * FROM vacancy ORDER by comid desc "
    cursor.execute(query)
    rows = cursor.fetchall()
    update(rows)

btna=Button(wrapper1,text="display all data",command=search_a,font=('Arial',10,'bold'))
btna.pack(side = LEFT, expand = True, fill = BOTH)

def get_treeview(trv):
    dat = []
    columns = []
    for column in trv["columns"]:
        columns.append(column)
    dat.append(columns)
    for item in trv.get_children():
        values = trv.item(item, 'values')
        dat.append(values)
    return dat

def create_excel(dat,filename):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    headers = ["Company ID","Company Name","Company Location","Position","Qualification","Age Limit","Vacancy Date"]  # Replace with your column names
# Add headers to the table
    dat.insert(0, headers)
    for row in dat:
        sheet.append(row)

    workbook.save(filename)
    print(f"Excel file '{filename}' created successfully.")


def excel():
    # Assuming your Treeview widget is called 'treeview'
    dat = get_treeview(trv)

    current_dir = filedialog.askdirectory(title="Select the current directory")
    if not current_dir:
        print("No directory selected.")
        return

    filename = filedialog.asksaveasfilename(
        title="Save Excel file",
        initialdir=current_dir,
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx")]
    )
    if not filename:
        print("No filename selected.")
        return

    create_excel(dat, filename)


excel_but = Button(wrapper1, text="Excel", command=excel, font=('Arial', 10, 'bold'))
excel_but.pack(side = LEFT, expand = True, fill = BOTH)



def upload_excel():
    # Open file dialog to select the Excel file
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    conn = sqlite3.connect("NASEOH.db")  # Replace "your_database.db" with your SQLite database file path
    cursor = conn.cursor()
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    for row in worksheet.iter_rows(values_only=True):
        query = "INSERT OR REPLACE INTO vacancy (comid,comname,location,position,quali,age,vacancy,vacdate) VALUES (?, ?, ?,?,?,?,?,?)"
        values = (row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7])
        cursor.execute(query, values)

    conn.commit()


upload = Button(wrapper1, text="Upload Excel", command=upload_excel, font=('Arial', 10, 'bold'))
upload.pack(side = LEFT, expand = True, fill = BOTH)

def nextpg():
    root.destroy()
    call(["python", "home.py"])
homepg = Button(wrapper3, text="HOMEPAGE", command=nextpg, font=('Arial', 10, 'bold')).grid(row=4, column=3, padx=5,pady=3)

up_btn.grid(row=4, column=0, padx=5, pady=3)
add_btn.grid(row=4, column=1, padx=5, pady=3)
delete_btn.grid(row=4, column=2, padx=5, pady=3)

root.mainloop()