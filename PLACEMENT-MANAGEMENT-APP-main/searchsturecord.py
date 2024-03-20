from tkcalendar import *
from tkinter import *
from tkinter import messagebox
import sqlite3
from tkinter import filedialog
import fpdf
from tkcalendar import *
from reportlab.lib.pagesizes import *
from reportlab.platypus import SimpleDocTemplate, Table, PageBreak
from reportlab.lib import colors
from tkinter import ttk
import os
from subprocess import call
from tkinter import colorchooser
import re
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfgen.canvas import Canvas
from reportlab.platypus import Paragraph, Frame, KeepInFrame
from tkinter import ttk
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import openpyxl


def update(rows):
    trv.delete(*trv.get_children())
    for i in rows:
        trv.insert('', 'end', values=i)


def search14():
    q2 = q.get()
    # q1 = q3.get()
    query = "Select cid,name,age,gender,email,address,quali,disability,disabtype,contact,ic,iccourse,camp,camploc,remarks,placement,date FROM registration WHERE name LIKE'%" + q2 + "%'"
    cursor.execute(query)
    rows = cursor.fetchall()
    update(rows)


def search15():
    q2 = q1.get()
    # q1 = q3.get()
    query = "Select cid,name,age,gender,email,address,quali,disability,disabtype,contact,ic,iccourse,camp,camploc,remarks,placement,date FROM registration WHERE cid LIKE'%" + q2 + "%'"
    cursor.execute(query)
    rows = cursor.fetchall()
    update(rows)


def search16():
    q2 = q3.get()
    # q1 = q3.get()
    query = "Select cid,name,age,gender,email,address,quali,disability,disabtype,contact,ic,iccourse,camp,camploc,remarks,placement,date FROM registration WHERE quali LIKE'%" + q2 + "%'"
    cursor.execute(query)
    rows = cursor.fetchall()
    update(rows)


def search17():
    q2 = selected_disab.get()
    # q1 = q3.get()
    query = "Select cid,name,age,gender,email,address,quali,disability,disabtype,contact,ic,iccourse,camp,camploc,remarks,placement,date FROM registration WHERE disability LIKE'%" + q2 + "%'"
    cursor.execute(query)
    rows = cursor.fetchall()
    update(rows)

def search18():
    q2 = q5.get()
    # q1 = q3.get()
    query = "Select cid,name,age,gender,email,address,quali,disability,disabtype,contact,ic,iccourse,camp,camploc,remarks,placement,date FROM registration WHERE contact LIKE'%" + q2 + "%'"
    cursor.execute(query)
    rows = cursor.fetchall()
    update(rows)

def search19():
    q2 = q6.get()
    q49 = q50.get()
    query = f"Select cid,name,age,gender,email,address,quali,disability,disabtype,contact,ic,iccourse,camp,camploc,remarks,placement,date FROM registration WHERE age BETWEEN {q2} AND {q49}"
    cursor.execute(query)
    rows = cursor.fetchall()
    cursor.execute(query)
    update(rows)

def search20():
    q2 = selected_placement.get()
    query = "Select cid,name,age,gender,email,address,quali,disability,disabtype,contact,ic,iccourse,camp,camploc,remarks,placement,date FROM registration WHERE placement = '" + q2 + "'" 
    
    cursor.execute(query)
    rows = cursor.fetchall()
    update(rows)


def search23():
    q2 = q8.get()
    query = "Select cid,name,age,gender,email,address,quali,disability,disabtype,contact,ic,iccourse,camp,camploc,remarks,placement,date FROM registration WHERE camploc LIKE'%" + q2 + "%'"
    cursor.execute(query)
    rows = cursor.fetchall()
    update(rows)

def search24():
    q2 = selected_iccourse.get()
    query = "Select cid,name,age,gender,email,address,quali,disability,disabtype,contact,ic,iccourse,camp,camploc,remarks,placement,date FROM registration WHERE iccourse LIKE'%" + q2 + "%'"
    cursor.execute(query)
    rows = cursor.fetchall()
    update(rows)


def search25():
    q2 = q9.get()
    query = "Select cid,name,age,gender,email,address,quali,disability,disabtype,contact,ic,iccourse,camp,camploc,remarks,placement,date FROM registration WHERE date LIKE'%" + q2 + "%'"
    cursor.execute(query)
    rows = cursor.fetchall()
    update(rows)

def clear():
    query = "SELECT cid,name,age,gender FROM registration"
    cursor.execute(query)
    rows = cursor.fetchall()
    update(rows)


def getrow(event):
    rowid = trv.identify_row(event.y)  # not really required
    item = trv.item(trv.focus())
    t1.set(item['values'][0])
    t2.set(item['values'][1])
    t3.set(item['values'][2])
    t4.set(item['values'][3])
    t5.set(item['values'][4])
    t6.set(item['values'][5])
    t7.set(item['values'][6])
    t8.set(item['values'][7])
    t9.set(item['values'][8])
    t10.set(item['values'][9])
    t11.set(item['values'][10])
    t12.set(item['values'][11])
    t13.set(item['values'][12])
    t14.set(item['values'][13])
    t15.set(item['values'][14])
    t16.set(item['values'][15])
    t17.set(item['values'][16])


def update_candidate():
    cid = t1.get()
    cname = t2.get()
    age = t3.get()
    gender = t4.get()
    email = t5.get()
    address = t6.get()
    quali = t7.get()
    disability = t8.get()
    disabtype = t9.get()
    contact = t10.get()
    ic = t11.get()
    iccourse = t12.get()
    camp = t13.get()
    camploc = t14.get()
    remarks = t15.get()
    placement = t16.get()
    date = t17.get()
    if messagebox.askyesno("Confirm update", "Are you sure you want to update the data?"):
        query = "UPDATE registration SET name=?,age=?,gender=?,email=?,address=?,quali=?,disability=?,disabtype=?,contact=?,ic=?,iccourse=?,camp=?,camploc=?,remarks=?,placement=?,date=?  WHERE cid=" + cid
        cursor.execute(query, (
            cname, age, gender, email, address, quali, disability, disabtype, contact, ic, iccourse, camp, camploc,
            remarks,
            placement, date))
        query1 = "SELECT cid,name,age,gender,email,address,quali,disability,disabtype,contact,ic,iccourse,camp,camploc,remarks,placement,date FROM registration"
        cursor.execute(query1)
        conn.commit()
    else:
        return True


def pick_date(event):
    global cal, date_window

    date_window = Toplevel()
    date_window.grab_set()
    date_window.title('Choose Date')
    date_window.geometry('250x250')
    cal = Calendar(date_window, selectmode='day', date_pattern='dd/mm/y')
    cal.place(x=0, y=0)

    sub_btn = Button(date_window, text='Submit', command=grab_date)
    sub_btn.place(x=80, y=190)


def grab_date():
    ent25.delete(0, END)
    ent25.insert(0, cal.get_date())
    date_window.destroy()


def pick_date1(event):
    global cal1, date_window1

    date_window1 = Toplevel()
    date_window1.grab_set()
    date_window1.title('Choose Date')
    date_window1.geometry('250x250')
    cal1 = Calendar(date_window1, selectmode='day', date_pattern='dd/mm/y')
    cal1.place(x=0, y=0)

    sub_btn = Button(date_window1, text='Submit', command=grab_date1)
    sub_btn.place(x=80, y=190)


def grab_date1():
    ent26.delete(0, END)
    ent26.insert(0, cal1.get_date())
    date_window1.destroy()


def convert_to_pdf():
    cid = t1.get()
    name = t2.get()
    age = t3.get()
    gender = t4.get()
    email = t5.get()
    address = t6.get()
    quali = t7.get()
    disability = t8.get()
    disabtype = t9.get()
    contact = t10.get()
    ic = t11.get()
    iccourse = t12.get()
    camp = t13.get()
    camploc = t14.get()
    remarks = t15.get()
    placement = t16.get()
    date = t17.get()

    print("NAME " + name)
    print("ID " + cid)
    print("GENDER " + gender)
    print("age" + age)
    print("EMAIL" + email)
    print("Contact" + contact)
    print("Location" + address)
    print("qualification" + quali)
    print("disability" + disability)
    print("type" + disabtype)
    print("Internal Candidate" + ic)
    print("Internal Candidate Course" + iccourse)
    print("Camp" + camp)
    print("Camp Loc" + camploc)
    print("Remark" + remarks)
    print("placement" + placement)
    print('date' + date)
    pdf = fpdf.FPDF()
    pdf.add_page()

    pdf.image("naseoh logo.jpg", 18, 10, w=200)

    pdf.set_font('Arial', size=20)
    pdf.text(58, 55, txt="CANDIDATE REGISTRATION FORM")

    pdf.set_font("Arial", size=12)

    pdf.text(30, 90, txt="Candidate ID:")
    pdf.text(60, 90, txt=cid)

    pdf.text(30, 80, txt="Date:")
    pdf.text(50, 80, txt=date)

    pdf.text(30, 100, txt="Candidate Name:")
    pdf.text(68, 100, txt=name)

    pdf.text(30, 110, txt="Address:")
    pdf.text(50, 110, txt=address)

    pdf.text(30, 140, txt="Email:")
    pdf.text(50, 140, txt=email)

    pdf.text(30, 150, txt="Gender:")
    pdf.text(50, 150, txt=gender)

    pdf.text(110, 150, txt="Age:")
    pdf.text(130, 150, txt=age)

    pdf.text(30, 160, txt="Contact:")
    pdf.text(50, 160, txt=contact)

    pdf.text(110, 160, txt="Qualification:")
    pdf.text(142, 160, txt=quali)

    pdf.text(30, 170, txt="Disability Type:")
    pdf.text(63, 170, txt=disabtype)

    pdf.text(110, 170, txt="Disability:")
    pdf.text(138, 170, txt=disability)

    pdf.text(30, 180, txt="Remarks:")
    pdf.text(55, 180, txt=remarks)

    pdf.text(110, 180, txt="Placement:")
    pdf.text(140, 180, txt=placement)

    pdf.text(30, 190, txt="Internal Candidate:")
    pdf.text(70, 190, txt=ic)

    pdf.text(30, 200, txt="Internal Candidate Course:")
    pdf.text(90, 200, txt=iccourse)

    pdf.text(30, 210, txt="Camp:")
    pdf.text(50, 210, txt=camp)

    pdf.text(30, 220, txt="Camp Location:")
    pdf.text(70, 220, txt=camploc)

    pdf.set_font("Arial", size=10)
    pdf.text(60, 280, txt="Developed by VES Institute of Technology ")
    pdf.image("vesitlogo.png", 129, 272, w=10, h=10)

    filepath = filedialog.asksaveasfilename(defaultextension=".pdf")
    pdf.output(filepath)


def delete_candidate():
    cid = t1.get()
    if messagebox.askyesno("Confirm delete", "Are you sure you want to delete the data?"):

        query = "DELETE FROM registration WHERE cid=" + cid
        cursor.execute(query)
        query1 = "SELECT cid,name,age,gender,email,address,quali,disability,disabtype,contact,ic,iccourse,camp,camploc,remarks,placement,date FROM registration"
        cursor.execute(query1)
        conn.commit()
        # conn.close()
        clear()
    else:
        return True

    cid = t1.get()
    cname = t2.get()
    age = t3.get()
    gender = t4.get()
    email = t5.get()
    address = t6.get()
    quali = t7.get()
    disability = t8.get()
    disabtype = t9.get()
    contact = t10.get()
    ic = t11.get()
    iccourse = t12.get()
    camp = t13.get()
    camploc = t14.get()
    remarks = t15.get()
    placement = t16.get()
    date = t17.get()

conn = conn = sqlite3.connect("NASEOH.db")
cursor = conn.cursor()

root = Tk()
q10 = StringVar()
q9 = StringVar()
q8 = StringVar()
q7 = StringVar()
q6 = IntVar()
q4 = StringVar()
q5 = StringVar()
q1 = StringVar()
q3 = StringVar()
q = StringVar()
q50 = IntVar()
q49 = IntVar()
q2 = StringVar()
t1 = StringVar()
t2 = StringVar()
t3 = StringVar()
t4 = StringVar()
t5 = StringVar()
t6 = StringVar()
t7 = StringVar()
t8 = StringVar()
t9 = StringVar()
t10 = StringVar()
t11 = StringVar()
t12 = StringVar()
t13 = StringVar()
t14 = StringVar()
t15 = StringVar()
t16 = StringVar()
t17 = StringVar()

root.title("Student Records")
width = root.winfo_screenwidth()
height = root.winfo_screenheight()
root.geometry("%dx%d" % (width, height))

wrapper1 = LabelFrame(root, text="Student List", font=('Arial', 10, 'bold'))
wrapper2 = LabelFrame(root, text="Search", font=('Arial', 10, 'bold'))
wrapper3 = LabelFrame(root, text="Update data", font=('Arial', 10, 'bold'))
wrapper1.configure(background="white")
wrapper2.configure(background="white")
wrapper3.configure(background="white")

wrapper1.pack(fill="both", expand="yes", padx=10, pady=5)
wrapper2.pack(fill="both", expand="yes", padx=10, pady=5)
wrapper3.pack(fill="both", expand="yes", padx=5, pady=5)

style = ttk.Style()
style.theme_use('alt')
style.configure("TCombobox", fieldbackground="white", background="white")
style.configure('TScrollbar', background="#329965")
style.configure('CustomScroll.Horizontal.TScrollbar', arrowcolor='white', background="#329965")
style.configure('CustomScroll.Vertical.TScrollbar', arrowcolor='white', background="#329965")
x_scroll = ttk.Scrollbar(wrapper1, orient=HORIZONTAL)
x_scroll['style'] = "CustomScroll.Horizontal.TScrollbar"
x_scroll.pack(side=BOTTOM, fill=X)

y_scroll = ttk.Scrollbar(wrapper1, orient=VERTICAL)
y_scroll['style'] = "CustomScroll.Vertical.TScrollbar"
y_scroll.pack(side=RIGHT, fill=Y)

trv = ttk.Treeview(wrapper1, columns=(1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17), show="headings",
                   height="5", xscrollcommand=x_scroll.set, yscrollcommand=y_scroll.set)
trv.pack(side = TOP, expand = True, fill = BOTH)
trv.heading(1, text="Candidate ID")
trv.heading(2, text="Candidate name")
trv.heading(3, text="Age")
trv.heading(4, text="Gender")
trv.heading(5, text="Email")
trv.heading(6, text="Address")
trv.heading(7, text="Qualification")
trv.heading(8, text="Disabiity")
trv.heading(9, text="Disability Type")
trv.heading(10, text="Contact")
trv.heading(11, text="Internal Candidate")
trv.heading(12, text="Internal Candidate Course")
trv.heading(13, text="Camp")
trv.heading(14, text="Camp Location")
trv.heading(15, text="Remarks")
trv.heading(16, text="Placements")
trv.heading(17, text='Date')

# important-event=double click and calling get row function i.e assigning function to the event
trv.bind('<Double 1>', getrow)

query = "SELECT cid,name,age,gender,email,address,quali,disability,disabtype,contact,ic,iccourse,camp,camploc,remarks,placement,date FROM registration ORDER BY cid desc"
cursor.execute(query)
rows = cursor.fetchall()
update(rows)

x_scroll.config(command=trv.xview)
y_scroll.config(command=trv.yview)


def get_treeview_data(trv):
    data = []
    for item in trv.get_children():
        values = trv.item(item, 'values')
        data.append(values)
    return data


def create_pdf(data):


    current_dir = filedialog.askdirectory(title="Select the current directory")

    if not current_dir:
        print("No directory selected.")
        exit()

    if not os.path.isdir(current_dir):
        print("Invalid directory path.")
        exit()


    filename = filedialog.asksaveasfilename(
        title="Save PDF file",
        initialdir=current_dir,
        defaultextension=".pdf",
        filetypes=[("PDF Files", "*.pdf")]
    )

    if not filename:
        print("No filename selected.")
        exit()


    doc = SimpleDocTemplate(filename, pagesize=A0)

    elements = []

    headers = ["Candidate ID", "Candidate name", "Age", "Gender", "Email", "Address", "Qualification", "Disabiity",
               "Disabiity Type", "Contact", "Internal Candidate", "Internal Candidate Course", "Camp", "Camp Location",
               "Remarks", "Placement", 'Date']  # Replace with your column names

    data.insert(0, headers)

    table = Table(data, colWidths=None, rowHeights=75, splitByRow=1,
                  repeatRows=0, repeatCols=0, rowSplitRange=None, spaceBefore=None,
                  spaceAfter=None, cornerRadii=None)


    style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 18),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 16),
        ('HALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),
        ('BOX', (0, 0), (-1, -1), 2, colors.black),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 15),
    ]

    table.setStyle(style)
    elements.append(table)

    doc.build(elements)
    print("PDF created successfully.")


def pdf():
    data = get_treeview_data(trv)
    create_pdf(data)


lbl14 = Label(wrapper2, text='Search by name', font=('Times-Roman', 10, 'bold'))
lbl14.grid(row=0, column=0, padx=5, pady=3)
ent14 = Entry(wrapper2, textvariable=q, highlightthickness=1, highlightbackground="black", highlightcolor="black")
ent14.grid(row=0, column=1, padx=5, pady=3)
btn14 = Button(wrapper2, text="Search", command=search14, font=('Arial', 10, 'bold'))
btn14.grid(row=0, column=2, padx=5, pady=3)

lbl15 = Label(wrapper2, text='Search by id', font=('Arial', 10, 'bold'))
lbl15.grid(row=1, column=0, padx=5, pady=3)
ent15 = Entry(wrapper2, textvariable=q1, highlightthickness=1, highlightbackground="black", highlightcolor="black")
ent15.grid(row=1, column=1, padx=5, pady=3)
btn15 = Button(wrapper2, text="Search", command=search15, font=('Arial', 10, 'bold'))
btn15.grid(row=1, column=2, padx=5, pady=3)

lbl16 = Label(wrapper2, text='Search by qualification', font=('Arial', 10, 'bold'))
lbl16.grid(row=2, column=0, padx=5, pady=3)
ent16 = Entry(wrapper2, textvariable=q3, highlightthickness=1, highlightbackground="black", highlightcolor="black")
ent16.grid(row=2, column=1, padx=5, pady=3)
btn16 = Button(wrapper2, text="Search", command=search16, font=('Arial', 10, 'bold'))
btn16.grid(row=2, column=2, padx=5, pady=3)

lbl17 = Label(wrapper2, text='Search by disability', font=('Arial', 10, 'bold'))
lbl17.grid(row=0, column=3, padx=5, pady=3)
selected_disab = StringVar(root, value="select type")
disab_menu = ttk.Combobox(wrapper2, textvariable=selected_disab, width=14, font=('Arial', 10), state='readonly')
disab_menu['values'] = ('OH', 'VI', 'ID', 'HI', 'SHI', 'MR', 'CP', 'OTHERS')
disab_menu.grid(row=0, column=4)
btn17 = Button(wrapper2, text="Search", command=search17, font=('Arial', 10, 'bold'))
btn17.grid(row=0, column=5, padx=5, pady=3)

lbl18 = Label(wrapper2, text='Search by contact', font=('Arial', 10, 'bold'))
lbl18.grid(row=1, column=3, padx=5, pady=3)
ent18 = Entry(wrapper2, textvariable=q5, highlightthickness=1, highlightbackground="black", highlightcolor="black")
ent18.grid(row=1, column=4, padx=5, pady=3)
btn18 = Button(wrapper2, text="Search", command=search18, font=('Arial', 10, 'bold'))
btn18.grid(row=1, column=5, padx=5, pady=16)

lbl19 = Label(wrapper2, text='Search by Age', font=('Arial', 10, 'bold'))
lbl19.grid(row=3, column=3, padx=5, pady=3)
lbl50 = Label(wrapper2, text='to')
lbl50.grid(row=3, column=5, padx=5, pady=3)
ent19 = Entry(wrapper2, textvariable=q6, highlightthickness=1, highlightbackground="black", highlightcolor="black")
ent19.grid(row=3, column=4, padx=5, pady=3)
ent50 = Entry(wrapper2, textvariable=q50, highlightthickness=1, highlightbackground="black", highlightcolor="black")
ent50.grid(row=3, column=6, padx=5, pady=3)
btn19 = Button(wrapper2, text="Search", command=search19, font=('Arial', 10, 'bold'))
btn19.grid(row=3, column=7, padx=5, pady=3)

lbl19 = Label(wrapper2, text='Search by Placement', font=('Arial', 10, 'bold'))
lbl19.grid(row=0, column=6, padx=5, pady=3)
selected_placement = StringVar(root, value="select type")
placement_menu = ttk.Combobox(wrapper2, textvariable=selected_placement, width=14, font=('Arial', 10), state='readonly')
placement_menu['values'] = ('Placed', 'Not Placed', 'Others')
placement_menu.grid(row=0, column=7)
btn19 = Button(wrapper2, text="Search", command=search20, font=('Arial', 10, 'bold'))
btn19.grid(row=0, column=8, padx=5, pady=3)

lbl23 = Label(wrapper2, text='Search by Camp Location', font=('Arial', 10, 'bold'))
lbl23.grid(row=1, column=6, padx=5, pady=3)
ent23 = Entry(wrapper2, textvariable=q8, highlightthickness=1, highlightbackground="black", highlightcolor="black")
ent23.grid(row=1, column=7, padx=5, pady=3)
btn23 = Button(wrapper2, text="Search", command=search23, font=('Arial', 10, 'bold'))
btn23.grid(row=1, column=8, padx=5, pady=3)

lbl24 = Label(wrapper2, text='Search by Internal Courses', font=('Arial', 10, 'bold'))
lbl24.grid(row=2, column=3, padx=5, pady=3)
selected_iccourse = StringVar(root, value="select course")
iccourse_menu = ttk.Combobox(wrapper2, textvariable=selected_iccourse, width=14, font=('Arial', 10), state='readonly')
iccourse_menu['values'] = (
'Baking', 'Gardening', 'Pottery', 'Welding', 'Assembly', 'Tailoring', 'Computer Skills', 'Data Processing',
'None')
iccourse_menu.grid(row=2, column=4)
btn24 = Button(wrapper2, text="Search", command=search24, font=('Arial', 10, 'bold'))
btn24.grid(row=2, column=5, padx=5, pady=3)

lbl26 = Label(wrapper2, text='Date', font=('Arial', 10, 'bold'))
lbl26.grid(row=3, column=0, padx=5, pady=3)
ent26 = Entry(wrapper2, textvariable=q9, highlightthickness=1, highlightbackground="black", highlightcolor="black")
ent26.grid(row=3, column=1, padx=5, pady=3)
ent26.insert(0, "dd/mm/yyyy")
ent26.bind("<1>", pick_date1)
btn26 = Button(wrapper2, text="Search", command=search25, font=('Arial', 10, 'bold'))
btn26.grid(row=3, column=2, padx=5, pady=3)

# User data section
lbl1 = Label(wrapper3, text='Candidate ID', font=('Arial', 10, 'bold'))
lbl1.grid(row=0, column=0, padx=5, pady=3)
ent1 = Entry(wrapper3, textvariable=t1, highlightthickness=1, highlightbackground="black", highlightcolor="black")
ent1.grid(row=0, column=1, padx=5, pady=3)

lbl2 = Label(wrapper3, text='Candidate name', font=('Arial', 10, 'bold'))
lbl2.grid(row=0, column=2, padx=5, pady=3)
ent2 = Entry(wrapper3, textvariable=t2, highlightthickness=1, highlightbackground="black", highlightcolor="black")
ent2.grid(row=0, column=3, padx=5, pady=3)

lbl3 = Label(wrapper3, text='Age', font=('Arial', 10, 'bold'))
lbl3.grid(row=1, column=0, padx=5, pady=3)
ent3 = Entry(wrapper3, textvariable=t3, highlightthickness=1, highlightbackground="black", highlightcolor="black")
ent3.grid(row=1, column=1, padx=5, pady=3)

lbl4 = Label(wrapper3, text='Gender', font=('Arial', 10, 'bold'))
lbl4.grid(row=1, column=2, padx=5, pady=3)
combo4 = ttk.Combobox(wrapper3, textvariable=t4, width=14, font=('Arial', 10), state='readonly')
combo4['values'] = ('Male', 'Female', 'Others')
combo4.grid(row=1, column=3, padx=5, pady=3)

lbl5 = Label(wrapper3, text='Email', font=('Arial', 10, 'bold'))
lbl5.grid(row=0, column=4, padx=5, pady=3)
ent5 = Entry(wrapper3, textvariable=t5, highlightthickness=1, highlightbackground="black", highlightcolor="black")
ent5.grid(row=0, column=5, padx=5, pady=3)

lbl6 = Label(wrapper3, text='Address', font=('Arial', 10, 'bold'))
lbl6.grid(row=1, column=4, padx=5, pady=3)
ent6 = Entry(wrapper3, textvariable=t6, highlightthickness=1, highlightbackground="black", highlightcolor="black")
ent6.grid(row=1, column=5, padx=5, pady=3)

lbl7 = Label(wrapper3, text='Qualification', font=('Arial', 10, 'bold'))
lbl7.grid(row=2, column=0, padx=5, pady=3)
ent7 = Entry(wrapper3, textvariable=t7, highlightthickness=1, highlightbackground="black", highlightcolor="black")
ent7.grid(row=2, column=1, padx=5, pady=3)

lbl8 = Label(wrapper3, text='Disability', font=('Arial', 10, 'bold'))
lbl8.grid(row=2, column=2, padx=5, pady=3)
combo8 = ttk.Combobox(wrapper3, textvariable=t8, width=14, font=('Arial', 10), state='readonly')
combo8['values'] = ('OH', 'VI', 'ID', 'HI', 'SHI', 'MR', 'CP', 'OTHERS')
combo8.grid(row=2, column=3, padx=5, pady=3)

lbl9 = Label(wrapper3, text='Disability Type', font=('Arial', 10, 'bold'))
lbl9.grid(row=2, column=4, padx=5, pady=3)
ent9 = Entry(wrapper3, textvariable=t9, highlightthickness=1, highlightbackground="black", highlightcolor="black")
ent9.grid(row=2, column=5, padx=5, pady=3)

lbl10 = Label(wrapper3, text='Contact', font=('Arial', 10, 'bold'))
lbl10.grid(row=3, column=0, padx=5, pady=3)
ent10 = Entry(wrapper3, textvariable=t10, highlightthickness=1, highlightbackground="black", highlightcolor="black")
ent10.grid(row=3, column=1, padx=5, pady=3)

lbl11 = Label(wrapper3, text='Internal Candidate', font=('Arial', 10, 'bold'))
lbl11.grid(row=3, column=2, padx=5, pady=3)
combo11 = ttk.Combobox(wrapper3, textvariable=t11, width=14, font=('Arial', 10), state='readonly')
combo11['values'] = ('Yes', 'No')
combo11.grid(row=3, column=3, padx=5, pady=3)

lbl12 = Label(wrapper3, text='Camp', font=('Arial', 10, 'bold'))
lbl12.grid(row=4, column=2, padx=5, pady=3)
combo12 = ttk.Combobox(wrapper3, textvariable=t13, width=14, font=('Arial', 10), state='readonly')
combo12['values'] = ('Yes', 'No')
combo12.grid(row=4, column=3, padx=5, pady=3)

lbl13 = Label(wrapper3, text='Remarks', font=('Arial', 10, 'bold'))
lbl13.grid(row=4, column=0, padx=5, pady=3)
ent13 = Entry(wrapper3, textvariable=t15, highlightthickness=1, highlightbackground="black", highlightcolor="black")
ent13.grid(row=4, column=1, padx=5, pady=3)

lbl20 = Label(wrapper3, text='Internal Candidate Course', font=('Arial', 10, 'bold'))
lbl20.grid(row=3, column=4, padx=5, pady=3)
combo20 = ttk.Combobox(wrapper3, textvariable=t12, width=14, font=('Arial', 10), state='readonly')
combo20['values'] = (
'Baking', 'Gardening', 'Pottery', 'Welding', 'Assembly', 'Tailoring', 'Computer Skills', 'Data Processing',
'None')
combo20.grid(row=3, column=5, padx=5, pady=3)

lbl21 = Label(wrapper3, text='Camp Location', font=('Arial', 10, 'bold'))
lbl21.grid(row=4, column=4, padx=5, pady=3)
ent21 = Entry(wrapper3, textvariable=t14, highlightthickness=1, highlightbackground="black", highlightcolor="black")
ent21.grid(row=4, column=5, padx=5, pady=3)

lbl22 = Label(wrapper3, text='Placement', font=('Arial', 10, 'bold'))
lbl22.grid(row=0, column=6, padx=5, pady=3)
combo22 = ttk.Combobox(wrapper3, textvariable=t16, width=14, font=('Arial', 10), state='readonly')
combo22['values'] = ('Placed', 'Not Placed', 'Others')
combo22.grid(row=0, column=7, padx=5, pady=3)

lbl25 = Label(wrapper3, text='Date', font=('Arial', 10, 'bold'))
lbl25.grid(row=1, column=6, padx=5, pady=3)
ent25 = Entry(wrapper3, textvariable=t17, highlightthickness=1, highlightbackground="black", highlightcolor="black")
ent25.grid(row=1, column=7, padx=5, pady=3)
ent25.insert(0, "dd/mm/yyyy")
ent25.bind("<1>", pick_date)


def reset():
    for widget in root.winfo_children():
        if isinstance(widget, Entry):
            widget.delete(0, 'end')
        if isinstance(widget, ttk.Combobox):
            widget.delete(0, 'end')


up_btn = Button(wrapper3, text="Update", command=update_candidate, font=('Arial', 10, 'bold'))
delete_btn = Button(wrapper3, text="Delete Data", command=delete_candidate, font=('Arial', 10, 'bold'))
pdf_button = Button(wrapper3, text="Convert to PDF", command=convert_to_pdf, font=('Arial', 10, 'bold'))

up_btn.grid(row=3, column=6, padx=5, pady=3)
delete_btn.grid(row=3, column=7, padx=5, pady=3)
pdf_button.grid(row=4, column=6, padx=5, pady=3)


pdf_but = Button(wrapper1, text="PDF", command=pdf, font=('Arial', 10, 'bold'))
pdf_but.pack(side = LEFT, expand = True, fill = BOTH)




def search_a():
    query = "Select cid,name,age,gender,email,address,quali,disability,disabtype,contact,ic,iccourse,camp,camploc,remarks,placement,date FROM registration ORDER by cid desc "
    cursor.execute(query)
    rows = cursor.fetchall()
    update(rows)


btna = Button(wrapper1, text="display all data", command=search_a, font=('Arial', 10, 'bold'))
btna.pack(side = LEFT, expand = True, fill = BOTH)


def get_treeview(trv):
    dat = []
    columns = []
    for column in trv["columns"]:
        columns.append(column)
    dat.append(columns)
    for item in trv.get_children():
        values = trv.item(item, 'values')
        dat.append(values)
    return dat


def create_excel(dat,filename):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    headers = ["Candidate ID", "Candidate name", "Age", "Gender", "Email", "Address", "Qualification", "Disability",
               "Disability Type", "Contact", "Internal Candidate", "Internal Candidate Course", "Camp", "Camp Location",
               "Remarks", "Placement", 'Date']  # Replace with your column names

    dat.insert(0, headers)
    for row in dat:
        sheet.append(row)

    workbook.save(filename)
    print(f"Excel file '{filename}' created successfully.")


def excel():

    dat = get_treeview(trv)

    current_dir = filedialog.askdirectory(title="Select the current directory")
    if not current_dir:
        print("No directory selected.")
        return

    filename = filedialog.asksaveasfilename(
        title="Save Excel file",
        initialdir=current_dir,
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx")]
    )
    if not filename:
        print("No filename selected.")
        return

    create_excel(dat, filename)


excel_but = Button(wrapper1, text="Excel", command=excel, font=('Arial', 10, 'bold'))
excel_but.pack(side = LEFT, expand = True, fill = BOTH)


def upload_excel():
    # Open file dialog to select the Excel file
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    # Connect to the SQLite database
    conn = sqlite3.connect("NASEOH.db")  # Replace "your_database.db" with your SQLite database file path
    cursor = conn.cursor()
    # Read data from the Excel file
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    # Iterate over rows and insert data into the database
    for row in worksheet.iter_rows(values_only=True):
        # Replace column_index with the appropriate index for each column in the row
        query = "INSERT OR REPLACE INTO registration (cid,name,age,gender,email,address,quali,disability,disabtype,contact,ic,iccourse,camp,camploc,remarks,placement,date) VALUES (?, ?, ?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
        values = (row[0],row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9],row[10],row[11],row[12],row[13],row[14],row[15],row[16])
        cursor.execute(query, values)

    conn.commit()


upload = Button(wrapper1, text="Upload Excel", command=upload_excel, font=('Arial', 10, 'bold'))
upload.pack(side = LEFT, expand = True, fill = BOTH)


def nextPage():
    root.destroy()
    call(["python", "home.py"])


b1 = Button(wrapper3, text="HOMEPAGE", command=nextPage, font=('Arial', 10, 'bold'))
b1.grid(row=4, column=7, padx=5, pady=3)

root.mainloop()